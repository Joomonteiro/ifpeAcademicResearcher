from os import link
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.firefox.options import Options
import time
from datetime import datetime
from openpyxl import load_workbook
from openpyxl import Workbook
import unidecode
import json
from elasticsearch import Elasticsearch
import unidecode
from pegaDados import ColetaDadosDoBancoEMandaProGCollector



es = Elasticsearch(HOST="http://localhost", PORT=9200)
es = Elasticsearch()


def transformaemminusculosemacentoeespaco(original):

    processamento_1 = unidecode.unidecode(original)

    processamento_2 = processamento_1.replace(" ", "")

    processamento_3 = processamento_2.lower()

    return processamento_3


def procuraDocentePorNome(nomeantigo):
    possivelmatch = 0
    # option = Options()
    # option.headless = True
    # o = webdriver.ChromeOptions()
    o = webdriver.FirefoxOptions()
    # o.binary_location = "C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    o.headless = True

    # driver = webdriver.Chrome(executable_path=r'./chromedriver.exe', options=o)
    driver = webdriver.Firefox(executable_path=r'./geckodriver.exe', options=o)
    nomeantigo = nomeantigo
    nomecompara = transformaemminusculosemacentoeespaco(nomeantigo)
    # #print("Nome em lowercase para comparação:")
    # #print(nomecompara)
    nome = parserNomeParaUrlDoGoogleScholar(nomeantigo)
    # #print("Nome após parser para utilização na URL:")
    # #print(nome)

    driver.get(
        "https://scholar.google.com.br/citations?hl=pt-BR&view_op=search_authors&mauthors="+nome+"&btnG=")
    # driver.get("http://sisbibliotec.com.br/a.html")
    # elem = driver.find_element_by_name('arroz')
    # elem.clear()
    # elem.send_keys('username')
    encontrado = 0
    encontradoGcollector = False
    # time.sleep(10)
    lista_elementos = driver.find_elements_by_class_name('gs_ai_t')
    # time.sleep(10)
    for i in lista_elementos:
        html = i.get_attribute("innerHTML")
        nomenovo = i.find_element_by_class_name('gs_hlt').text
        nomenovo = str(nomenovo)
        nomenovo = transformaemminusculosemacentoeespaco(nomenovo)
        # #print("Nome encontrado no resultado do google scholar após parser para comparação:")
        # #print(nomenovo)
        if 'Instituto Federal' in html or 'ifpe' in html:
            # html = i.get_attribute("innerHTML")
            # #print("Encontrado!")
            
            link = i.find_element_by_css_selector('a').get_attribute('href')
            # #print(link)
            encontradoGcollector = True
            encontrado = 1
            break
        elif (nomecompara == nomenovo):
            link = i.find_element_by_css_selector('a').get_attribute('href')
            # #print("encontrado nome completo! possível match")
            # #print(link)
            linkmatching = link
            possivelmatch = 1
            encontrado = 1
            break
    
    if encontrado == 1:
        # #print("redirecionando para o escavador")
        driver.get("https://www.escavador.com/busca?qo=p&q=" +
                   nome+"&f%5Bpt%5D%5B%5D=curriculo")
        lista_elementos = driver.find_elements_by_class_name('item')
        #print(lista_elementos)
        for i in lista_elementos:
            html = i.get_attribute("innerHTML")

            # nomeOcid = nomeOcidList[1]
            # nomeOcid = str(nomeOcid)
            # nomeOcid = transformaemminusculosemacentoeespaco(nomeOcid)
            #print(nomeOcid)

            #print(html)
            if 'Instituto Federal' in html or 'ifpe' in html:
                # html = i.get_attribute("innerHTML")
                # #print("Encontrado!")
                link1 = i.find_element_by_css_selector(
                    'a').get_attribute('href')
                # #print(link1)
                encontrado = 1
                if (encontradoGcollector):
                    coletaDadosDeProducoes(link,link1,nomecompara)

                if(not encontradoGcollector):
                    verificaSimilaridadeDeProducoes(
                        possivelmatch, link1, linkmatching, nomecompara)

                break
            # elif (nomecompara=)

    if encontrado == 1:
        nome = nome.replace("+", "%20")
        # #print("redirecionando para o Orcid")
        driver.get("https://orcid.org/orcid-search/search?searchQuery="+nome)
        time.sleep(1)
        lista_elementos = driver.find_elements_by_class_name(
            'ng-star-inserted')
        #print(lista_elementos)
        for i in lista_elementos:

            html = i.get_attribute("innerHTML")
            # nomeOcidList = i.find_elements_by_tag_name('td')
            #print(nomeOcidList[1].text)
            #print(html)
            if 'Instituto Federal' in html or 'ifpe' in html:
                # html = i.get_attribute("innerHTML")
                link1 = i.find_element_by_css_selector(
                    'a').get_attribute('href')
                # #print("Encontrado!")
                # #print(link1)
                encontrado = 1
                break
    elif(encontrado == 0):
        # #print(nomecompara+" "+nomenovo)
        print("")
        # #print("Não encontrado em nenhuma das bases")
    driver.quit()


def parserNomeParaUrlDoGoogleScholar(nomeantigo):
    nome = nomeantigo.strip()
    nome = nome.replace(" ", "+")
    return nome


def verificaSimilaridadeDeProducoes(possivelmatch, link1, linkmatching, nomecompara):
    if possivelmatch == 1:
        similaridade1, todasAsProducoes, producoesEcollector = verificaMatching(
            link1, linkmatching)
        if similaridade1 == 1:
            # #print(
                # "match confirmado, armazenando dados do match e #printando os dados de produções de ambos")
            # #print("Produções no google scholar:")
            # #print(todasAsProducoes)
            saveProducoes(todasAsProducoes, nomecompara)

            # #print("Produções no escavador:")
            # #print(producoesEcollector)
    

def coletaDadosDeProducoes( link, link1, nomecompara):
    
    todasAsProducoes, producoesEcollector = coletadadosDoGcollectorEEcollector(link, link1)
        
    # #print("salvando produções do google scholar")
    # #print("Produções no google scholar:")
    # #print(todasAsProducoes)
    saveProducoes(todasAsProducoes, nomecompara)

    # #print("Produções no escavador:")
    # #print(producoesEcollector)

def saveProducoes(todasAsProducoes, nomecompara):
    saveProducoesNoBancoElastic(todasAsProducoes, nomecompara)
    saveProducoesNaPlanilhaExel(todasAsProducoes, nomecompara)


def saveProducoesNoBancoElastic(todasAsProducoes, nomecompara):
    ids = 1
    for producao in todasAsProducoes:
        resultado = transformaemminusculosemacentoeespaco(
            nomecompara)
        # #print(resultado)
        res = es.index(
            index=resultado, doc_type="trabalhosacademicos", id=ids, body=producao)
        ids = ids + 1
        # #print(res['result'])
    ColetaDadosDoBancoEMandaProGCollector(resultado)


def saveProducoesNaPlanilhaExel(todasAsProducoes, nomecompara):
    wb = Workbook()
    ws = wb.active
    ws.title = nomecompara
    # #print("Titulo da planilha:")
    # #print(nomecompara)
    indexi = 0
    indexj = 0
    lista = []
    for producao in todasAsProducoes:
        autor = (producao.get('autor'))
        areas = (producao.get('areas'))
        titulo = (producao.get('título'))
        colaboradores = (producao.get('text1'))  
        localDePublicacao = (producao.get('text2'))
        citacoes = (producao.get('citações'))
        ano = (producao.get('Ano'))
        totalDeCitacoes = (producao.get('totaldecitacoes'))
        timestamp = (producao.get('timestamp'))
        lista.append([autor, areas, titulo, colaboradores, localDePublicacao, citacoes, ano, totalDeCitacoes,timestamp])
        # lista = [autor, areas, titulo, colaboradores, localDePublicacao, citacoes, ano, totalDeCitacoes,timestamp]
    #print(lista)
    maxLinhas = len(lista)  
    maxColunas = len(lista[0])   
    
        # linha = []
        # linha.append(lista)
        

    # linhaLista = len(lista)
    # #print(lista)
    # for i in range(17):
    #     for j in range(9):
    #         valor = str(lista[indexi][indexj])
            # #print(valor)
    #         indexj += 1
            # #print(indexj)
    #     indexj = 0
        # #print(indexi)
    #     indexi = indexi + 1
    for row in ws.iter_rows(min_row=1, max_col=maxColunas, max_row=maxLinhas):
        for cell in row:

            valor = str(lista[indexi][indexj])
            # #print(valor)
            # #print(cell)
            cell.value = valor
            # #print(cell.value)
            indexj += 1
            # #print(indexj)
        indexj = 0
        # #print(indexi)
        indexi = indexi + 1
        
    wb.save(nomecompara+'.xlsx')


def gcollector(link):
    option = Options()
    option.headless = True
    # o = webdriver.ChromeOptions()
    # o.binary_location = "C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"
    # o.headless = True

    # driver = webdriver.Chrome(executable_path=r'./chromedriver.exe', options=o)
    driver = webdriver.Firefox(executable_path=r'./geckodriver.exe', options=option)
    driver.get(link)
    # action = ActionChains(driver)

    c = True
    while(c):
        menuid = driver.find_element_by_id('gsc_bpf_more')
        menuativado = menuid.is_enabled()
        if(menuativado):
            menuid.click()
            time.sleep(1.5)
        if menuativado is False:
            c = False
    # menu = driver.find_element_by_xpath('//*[@id="gsc_bpf_more"]/span/span[1]')
    # menuid = driver.find_element_by_id('gsc_bpf_more')
    # #print(menuid.is_enabled())
    # prop = menu.get_property('disabled')
    # #print(prop)
    # menu.click()
    # time.sleep(2)
    # menu = driver.find_element_by_xpath('//*[@id="gsc_bpf_more"]/span/span[1]')
    # menu.click()
    # time.sleep(2)
    # menu = driver.find_element_by_xpath('//*[@id="gsc_bpf_more"]/span/span[1]')
    # menu.click()
    # time.sleep(2)
    # menu = driver.find_element_by_xpath('//*[@id="gsc_bpf_more"]/span/span[1]')
    # menu.click()
    # html = driver.page_source
    dados = driver.find_element_by_id("gs_bdy_ccl")
    html = dados.get_attribute("innerHTML")

    soup = BeautifulSoup(html, "html.parser")
    #print(soup)

    #print(soup)
    ind = 0
    ids = 2
    # 3 separating the data we want by class
    gsces = soup.find_all(class_="gsc_a_tr")
    texts = (soup.find_all(class_="gs_gray"))
    totaldecitacoes = (soup.find(class_="gsc_rsb_std"))
    totaldecitacoes = totaldecitacoes.text
    vetorareas = []
    stringareas = ""
    areas = (soup.find(id="gsc_prf_int"))
    #print(len(areas))
    for area in areas:
        area = area.text
        area = area + ", "
        stringareas = stringareas + area
        vetorareas.append(area)
    # #print(stringareas)

    autor = (soup.find(id="gsc_prf_in"))
    autor = autor.text

    all_texts1 = []
    all_texts2 = []
    all_productions = []

    for i in range(len(texts)):
        if(i % 2 == 0):
            text1 = ((texts[i]).text)
            all_texts1.append(text1)

    for i in range(len(texts)):

        if(i % 2 != 0):

            text2 = ((texts[i]).text)
            all_texts2.append(text2)
    titulos = []
    # 4 collecting the data we want and organizing it in dictionary form
    for gsc in gsces:
        #print(ids)
        titles = (gsc.find(class_="gsc_a_at").text)
        titulos.append(titles.lower())
        try:
            amount_of_citations = (gsc.find(class_="gsc_a_ac gs_ibl").text)
        except:
            amount_of_citations = '0'
        year_of_publication = (gsc.find(class_="gsc_a_h gsc_a_hc gs_ibl").text)

        all_productions.append({'autor': autor,
                                'areas': stringareas,
                                'título': titles,
                                'text1': all_texts1[ind],
                                'text2': all_texts2[ind],
                                'citações': amount_of_citations,
                                'Ano': year_of_publication,
                                'totaldecitacoes': totaldecitacoes,
                                'timestamp': datetime.now(),
                                })
        ind += 1
    #print(all_productions)
    #print(titulos)
    driver.quit()
    return titulos, all_productions


def ecollector(link):
    # 1 taking the response of requests
    res = requests.get(link)

    # 2 putting the encoding as "utf8" to collect Latin characters without being encoded
    res.encoding = "utf-8"

    # 3 parsing the soup
    soup = BeautifulSoup(res.text, 'html.parser')

    # 4 collecting the required data
    # resumes = soup.find(class_="box -flushHorizontal")
    # resumes1 = resumes.find('p').text
    boxes1 = soup.find(id="producoes")
    #print(boxes1)
    boxes = boxes1.find('ul')
    #print(boxes)
    boxes2 = boxes.find_all('p')
    #print(boxes2)
    formations = soup. find_all(
        id="formacao", class_="box -flush inline-edit-main-box")
    historic = soup.find_all(id="atuacao-profissional",
                             class_="box -flush inline-edit-main-box")

    # 5 extracting only the text inside the tags and organizing them within a dictionary
    all_productions = []
    # all_productions1 = []
    # ids=0

    producoes = []
    producoes1 = []
    for box in boxes2:
        # minha_string = minha_string.replace('\n', '')
        #print(minha_string)
        #print(box.text)
        #print(ids)
        string = box.text
        string = string.replace('\n', '')
        producoes1.append(string)
        producoes.append(string.lower())
    return producoes, boxes2, boxes, producoes1


def verificaMatching(link, linkmatching):
    linkmatching = linkmatching
    titulos, todasAsProducoes = gcollector(linkmatching)
    # #print("Titulos de produções do Google Scholar após parser para lowercase:")
    # #print(titulos)

    producoes, boxes2, boxes, producoesEcollector = ecollector(link)
    # #print(type(boxes))
    boxes = str(boxes)
    producoesul = boxes.lower()
    # #print("Produções do Escavador após parser para lowercase:")
    # #print(producoesul)
    # parserAbnt(boxes2)
    #print(type(producoes))
    # producoes = str(producoes)
    #print(type(producoes))
    #print(producoes)

    similaridade = 0
    for titulo in titulos:
        if titulo in producoesul:
            # #print("Titulo encontrado para o matching:")
            # #print(titulo)
            similaridade += 1
            break
    #print(similaridade)
    return similaridade, todasAsProducoes, producoesEcollector

def coletadadosDoGcollectorEEcollector(link, link1):
    link = link
    titulos, todasAsProducoes = gcollector(link)
    # #print("Titulos de produções do Google Scholar após parser para lowercase:")
    # #print(titulos)

    producoes, boxes2, boxes, producoesEcollector = ecollector(link1)
    # #print(type(boxes))
    boxes = str(boxes)
    producoesul = boxes.lower()
    # #print("Produções do Escavador após parser para lowercase:")
    # #print(producoesul)
    # parserAbnt(boxes2)
    #print(type(producoes))
    # producoes = str(producoes)
    #print(type(producoes))
    #print(producoes)

    # Aqui se fazia a comparação se não tivesse encontrado do gcollector
    # similaridade = 0
    # for titulo in titulos:
    #     if titulo in producoesul:
            #print("Titulo encontrado para o matching:")
            #print(titulo)
    #         similaridade += 1
    #         break
    #print(similaridade)
    return  todasAsProducoes, producoesEcollector

def coletarDadosDoBanco():
    return [{'autor': 'tiago'}]

def procuraDocentes(listaDeDocentes):
    listaDeDocentes = listaDeDocentes
    for docente in listaDeDocentes:
        nomeantigo = docente
        # #print("Agora procurando e coletando dados do docente:")
        # #print(nomeantigo)
        procuraDocentePorNome(nomeantigo)

# listaDeDocentes = ['Tiago Pessoa Ferreira de Lima','Tiago Brasileiro Araújo']
# nomeantigo = input("Digite o nome do docente:")
